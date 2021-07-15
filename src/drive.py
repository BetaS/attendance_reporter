from openpyxl import load_workbook
from datetime import date, time
from dateutil.relativedelta import relativedelta

import holidays

from src.user_db import UserModel


kr_holidays = holidays.Korea()


def make_data(path, year, month, user: UserModel):
    wb = load_workbook("./출근부 샘플.xlsx")
    sheet = wb.worksheets[0]

    start = date(year, month, 1)
    end = start + relativedelta(months=1, days=-1)

    sheet["B3"].value = f"{user.idx:07}"
    sheet["B4"].value = user.name
    sheet["E3"].value = start
    sheet["G3"].value = end

    for d, t in user.times.items():
        day = date(year, month, d)
        sheet[f"C{d+6}"] = "휴일" if day.isoweekday() > 5 or day in kr_holidays else ""

        for i in range(len(t)):
            sheet.cell(d+6, 8+i, t[i])

    wb.save(f"{path}/{user.idx}_{user.name}_출퇴근기록_{year}_{month:02}.xlsx")
